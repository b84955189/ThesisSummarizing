#!/usr/bin/env python3
# -*- encoding: utf-8 -*-
"""
@File    :   ExcelFunc.py    
@Contact :   lking@lking.icu
@Author :    Jason
@Date :      2022/4/28 19:36
@Description  Python version-3.10
TODO： 优化错误信息记录的方式
TODO： 记录非依据表中的多余评分记录
"""
from pathlib import Path

from openpyxl import load_workbook

from func import CommonTools
from model.EnumModels import SheetType

from model.ExceptionModels import InvalidRatingRecordException, DuplicatedRatingRecordException, \
    StudentRatingRecordIncompleteException
from model.RatingModels import CommentScoreModel, DebateScoreModel, TeacherScoreModel, OutputModel

# 每表 分数 固定从第 6 列 开始检索
SCORE_INDEX = 6
# 表 标题 数据行位置 固定为 2
HEADER_ROW = 2
# Excel 表名称
COMMENT_SCORE_SHEET_NAME = "评阅老师成绩"
DEBATE_SCORE_SHEET_NAME = "答辩成绩"
TEACHER_SCORE_SHEET_NAME = "指导老师成绩"


def get_workbook(file_path):
    return load_workbook(file_path)


def close_workbook(workbook):
    """
    关闭Excel工作簿对象
    @param workbook: Excel工作簿对象
    @return:
    """
    if workbook is not None:
        workbook.close()


def get_data_from_sheet(sheet, sheet_type):
    """
    读 成绩 记录
    @param sheet: 表对象
    @param sheet_type: 表类型 - 枚举
    @return: 表数据 - dict
    """
    models = {}
    error = []
    # 跳过 成绩表 表头
    for row in range(3, sheet.max_row + 1):
        model = None
        try:
            # 因 不同模板表之间存在固定相同的 列数据，所以预先将相同的列数据读取，以便后续进行区别化操作
            row_id = sheet.cell(row, 1).value
            row_student_number = sheet.cell(row, 2).value
            row_student_name = sheet.cell(row, 3).value
            row_major = sheet.cell(row, 4).value
            row_thesis_topic = sheet.cell(row, 5).value
            # 判断该评分记录是否有效，是否值得读入。
            # 这里主要通过判断是否有学号与姓名来判断
            # 如果为学号或姓名为空，跳过该评分记录
            if CommonTools.is_empty_or_none(row_student_number) or CommonTools.is_empty_or_none(
                    row_student_name):
                raise InvalidRatingRecordException(f"【{sheet}】表 的第【{row}】行数据无效（缺少学号、姓名或指导老师信息）！\n")
            # 分数
            # 每表分数固定从第 SCORE_INDEX 列 开始检索
            row_column = SCORE_INDEX
            scores = []
            # 表 标题 数据行位置 固定为 HEADER_ROW
            while "得分" in sheet.cell(HEADER_ROW, row_column).value:
                scores.append(
                    0 if sheet.cell(row, row_column).value is None else int(sheet.cell(row, row_column).value))
                row_column += 1
            match sheet_type:
                case SheetType.COMMENT_SCORES_SHEET:
                    model = CommentScoreModel()
                    # row_column+1 表示跳过 表的 ‘合计’ 列，因为合计分数由本程序读取分数计算得出
                    # 后续依次+1 即为后续固定列
                    model.review_teacher_name = sheet.cell(row, row_column + 1).value
                    model.review_teacher_work_number = sheet.cell(row, row_column + 2).value
                case SheetType.DEBATE_SCORES_SHEET:
                    model = DebateScoreModel()
                    # row_column+1 表示跳过 表的 ‘合计’ 列，因为合计分数由本程序读取分数计算得出
                    # 后续依次+1 即为后续固定列
                    model.debate_content = sheet.cell(row, row_column + 1).value
                    model.debate_group_leader_name = sheet.cell(row, row_column + 2).value
                    model.debate_group_secretary_name = sheet.cell(row, row_column + 3).value
                    model.debate_group_member = sheet.cell(row, row_column + 4).value
                case SheetType.TEACHER_SCORES_SHEET:
                    model = TeacherScoreModel()
                    # row_column+1 表示跳过 表的 ‘合计’ 列，因为合计分数由本程序读取分数计算得出
                    # 后续依次+1 即为后续固定列
                    model.guidance_teacher_name = sheet.cell(row, row_column + 1).value
                    model.guidance_teacher_work_number = sheet.cell(row, row_column + 2).value
                case _:
                    # 不做处理
                    pass

            model.id = row_id
            model.student_number = row_student_number
            model.student_name = row_student_name
            model.major = row_major
            model.thesis_topic = row_thesis_topic
            model.scores = scores
            if model.student_number not in models:
                models[model.student_number] = model
            else:
                raise DuplicatedRatingRecordException(f"学号:【{model.student_number}】姓名：【{model.student_name}】的评分记录在【{sheet}】表的第【{row}】行重复（已忽略该重复记录）。\n")
        # 记录错误信息
        except InvalidRatingRecordException as invalidRatingRecord:
            error.append(str(invalidRatingRecord))
        except DuplicatedRatingRecordException as duplicatedRatingRecordException:
            error.append(str(duplicatedRatingRecordException))
        finally:
            pass

    return models, error


def reorganization_data(debate_scores_and_error, comment_scores_and_error, teacher_scores_and_error):
    """
    重组数据 - 一个学生 对应 三个评分表
    循环 依据表 - 指导老师评分表
    @param debate_scores_and_error: 带有错误信息的 答辩评分表 数据
    @param comment_scores_and_error: 带有错误信息的 评阅评分表 数据
    @param teacher_scores_and_error: 带有错误信息的 指导老师评分表 数据
    @return:
    """
    error = []
    # 评分
    teacher_scores = teacher_scores_and_error[0]
    comment_scores = comment_scores_and_error[0]
    debate_scores = debate_scores_and_error[0]
    # 异常
    error.extend(teacher_scores_and_error[1])
    error.extend(comment_scores_and_error[1])
    error.extend(debate_scores_and_error[1])

    output_models = []
    for k, v in teacher_scores.items():
        # print(k, v)
        output_model = OutputModel()
        # 基本信息  - 这里不可能为空 ，因为读数据时已经经过筛选
        output_model.student_number = v.student_number
        output_model.student_name = v.student_name
        output_model.guidance_teacher_name = v.guidance_teacher_name
        # 评分 - 这里可能会出现 评分信息不完整的情况，所以判断一下   - 指导老师成绩为依据
        try:
            output_model.teacher_score_model = v
            if comment_scores.get(v.student_number) is None:
                raise StudentRatingRecordIncompleteException(
                    f"学号：【{v.student_number}】 姓名：【{v.student_name}】 缺少 【评阅老师成绩】 记录。\n")
            if debate_scores.get(v.student_number) is None:
                raise StudentRatingRecordIncompleteException(
                    f"学号：【{v.student_number}】 姓名：【{v.student_name}】 缺少 【答辩成绩】 记录。\n")
            output_model.comment_score_model = comment_scores[v.student_number]
            output_model.debate_score_model = debate_scores[v.student_number]
            output_models.append(output_model)
        except StudentRatingRecordIncompleteException as e:
            error.append(str(e))
        finally:
            pass

    return output_models, error


